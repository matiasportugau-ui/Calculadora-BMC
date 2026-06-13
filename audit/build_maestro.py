#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BMC Uruguay — Technical Audit MAESTRO generator.
Consolidates the 2026-06-07 exhaustive audit into:
  - BMC-MAESTRO-AUDIT.xlsx   (multi-tab; upload to Drive -> Google Sheet)
  - MAESTRO_data/*.csv       (one CSV per tab, importable)
  - dashboard.html           (visual summary)
  - audit-metadata.json      (machine-readable full report)

Read-only audit. No secret values, no client PII. Run: python3 audit/build_maestro.py
"""
import csv, json, os

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "MAESTRO_data")
os.makedirs(DATA, exist_ok=True)

AUDIT_DATE = "2026-06-07"
PRIOR_DATE = "2026-04-23"
LIVE_SHA   = "3153b111eea1979cbc3766346b51c7eac003420e"

# ---------------------------------------------------------------------------
# TAB 0 — OVERVIEW
# ---------------------------------------------------------------------------
overview = {
    "title": "BMC Uruguay — Auditoría Técnica MAESTRO",
    "org": "BMC Uruguay / METALOG SAS (RUT 120403430012)",
    "system": "Calculadora-BMC (quote engine + CRM + dashboards), v3.1.5",
    "repo": "matiasportugau-ui/calculadora-bmc",
    "audit_date": AUDIT_DATE,
    "builds_on": f"Auditoría previa {PRIOR_DATE} en .relevamiento/matriz/ (sin acceso a gh/gcloud/Drive). Esta corre CON esos accesos.",
    "live_prod": {
        "cloud_run_url": "https://panelin-calc-q74zutv7dq-uc.a.run.app",
        "health": "HTTP 200 — appEnv=production, hasSheets=true, hasTokens=true, mlTokenStoreOk=true",
        "deployed_git_sha": LIVE_SHA,
        "matches_local_main": True,
        "missing_config": ["ML_CLIENT_SECRET"],
        "version": "3.1.5",
    },
    "headline_numbers": {
        "tracked_files": 2413,
        "remote_branches": 53,
        "open_prs": 45,
        "api_endpoints": 226,
        "public_endpoints": 70,
        "google_sheets_audited": 8,
        "sibling_repos_in_org": 24,
        "security_findings": 15,
        "critical_findings": 3,
    },
    "deltas_vs_prior": [
        "Ramas: 84 -> 53 (limpieza ocurrida).",
        "PRs abiertas: ~20 -> 45 (creció; muchas son fix de deploy/CI y test-coverage de bots).",
        "Sheets: antes 'duda abierta' -> ahora 8/8 leídas en vivo via Drive MCP.",
        "Repos hermanos: antes no enumerables -> ahora 24 listados (solo 2 activos).",
        "Prod Cloud Run: antes no verificable -> ahora confirmado vivo y = main.",
        "P3 (schema CRM): antes incógnita -> CRM_Operativo confirmado en vivo via /health.",
        "PII: prior dijo 'sin exposición' -> ESTA encuentra CRM con 297 leads commiteado (.accessible-base/).",
    ],
}

# ---------------------------------------------------------------------------
# TAB 1 — INFRAESTRUCTURA
# ---------------------------------------------------------------------------
infra_cols = ["Capa","Componente","URL / ID","Auth / Acceso","Estado (vivo)","Notas"]
infra = [
    ["Frontend","Vercel SPA (Vite)","https://calculadora-bmc.vercel.app","Pública + Google OAuth","ACTIVO","Build vite -> dist/. Rewrites /api,/calc,/auth,/sync -> Cloud Run."],
    ["API","Cloud Run 'panelin-calc' (us-central1)","https://panelin-calc-q74zutv7dq-uc.a.run.app","JWT / API_AUTH_TOKEN / HMAC","ACTIVO (HTTP 200)","gitSha en vivo 3153b11 == main. Express 5, Node 24, server/Dockerfile."],
    ["API (URL alt)","Mismo servicio, URL por nº de proyecto","https://panelin-calc-642127786762.us-central1.run.app","idem","ACTIVO (alias)","Misma service; 642127786762 = nº proyecto. Canónica = q74 (public_base_url + vercel.json)."],
    ["GCP Project","chatbot-bmc-live / us-central1","Artifact Registry: cloud-run-repo/","WIF + deploy SA","ACTIVO","Phase 0 SA hardening 2026-06-01: secrets en Secret Manager."],
    ["Imagen API (canónica)","server/Dockerfile (node:24-alpine, :8080)","deploy-calc-api.yml + cloudbuild-api.yaml","-","CANÓNICA","API-only + Chromium para PDF. ES la imagen de prod."],
    ["Imagen full-stack (alt)","Dockerfile.bmc-dashboard","cloudbuild.yaml + scripts/deploy-cloud-run.sh","-","ALTERNATIVA/MANUAL","SPA+API+Debian Chromium. NO es el path CI activo."],
    ["Imagen frontend (alt)","Dockerfile (root -> nginx)","cloudbuild-frontend.yaml -> panelin-calc-web","-","ALTERNATIVA","nginx static. Servicio panelin-calc-web (legacy/fallback)."],
    ["Datos: Sheets","Google Sheets (8 libros) via service account","bmc-dashboard-sheets@chatbot-bmc-live.iam","Service account JSON (Secret Manager)","ACTIVO","Fuente CRM/Finanzas/Precios. Ver tab Sheets Analysis."],
    ["Datos: Postgres","Supabase (htnwozvopveibwppyjhg, us-east-1)","DATABASE_URL (Secret Manager)","Connection string","ACTIVO","WA Cockpit, Transportista, TraKtiMe, Identity, Tasks, RAG pgvector."],
    ["Datos: GCS","Bucket bmc-cotizaciones","allUsers:objectViewer (PDFs)","Service account","ACTIVO","PDFs de cotización, evidencias transportista, facturas traktime."],
    ["Apps Script","5 .gs en docs/bmc-dashboard-modernization/","Container-bound a workbooks","Manual","SIN PIPELINE","Code.gs, CalendarioRecordatorio, PagosPendientes, StockAlertas, VentasConsolidar. Sin deploy automatizado (P5)."],
    ["Integración: MercadoLibre","OAuth + API (MLU)","ML_CLIENT_ID 742811153438318","OAuth","PARCIAL","ML_CLIENT_SECRET FALTA en prod (/health missingConfig)."],
    ["Integración: WhatsApp","WhatsApp Cloud API + webhook HMAC","/webhooks/whatsapp","App secret (HMAC)","CONFIG","WA Cockpit en Postgres."],
    ["Integración: Shopify","OAuth + webhooks + shop-chat-agent/","/webhooks/shopify","HMAC","SUB-APP","shop-chat-agent tiene Dockerfile propio (Node 18) — posible servicio separado."],
    ["Integración: LLM","Anthropic (claude-opus-4-7) + OpenAI fallback + Vercel AI Gateway","-","API keys (Secret Manager)","ACTIVO","agentCore chain; RAG pgvector."],
    ["CI/CD","GitHub Actions (9 workflows)",".github/workflows/","-","ACTIVO","ci.yml, deploy-calc-api.yml, deploy-vercel.yml, smoke-prod-scheduled, knowledge-antenna."],
    ["Secret mgmt","Doppler (bmc-backend/prd) + GCP Secret Manager","-","-","ACTIVO","Phase 0: API_AUTH_TOKEN, ML_CLIENT_SECRET, ANTHROPIC/GEMINI/GROK keys, TOKEN_ENCRYPTION_KEY."],
]

# ---------------------------------------------------------------------------
# TAB 2 — SHEETS ANALYSIS
# ---------------------------------------------------------------------------
sheets_cols = ["#","Sheet (label)","Spreadsheet ID","Propósito real","Tabs / estructura","Problemas de calidad","Estado"]
sheets = [
 ["1","2.0 - Ventas Dashboard","1KFNKWLQmBHj_v8BZJDzLklUtUPbNssbYEsWcmc0KPQA",
  "Tracker de pedidos/ventas activo (MONTO/COSTO/GANANCIAS/SALDOS)",
  "Tab instrucciones (Apps Script copy) + pipeline principal ~30 filas (26 cols) + bloques NO_HEADER scratch",
  "Header doble/merged (ID.Pedido x2); notas de pago libres mezcladas en columnas estructuradas","AMARILLO"],
 ["2","MATRIZ de COSTOS","1oDMkBgWxX7cu7TpSvuO30tCTUWl68IBDhC4cQTP79Xo",
  "MASTER de precios/costos (fuente de /api/actualizar-precios-calculadora; crítico smoke:prod)",
  "Bloques por familia (ISOROOF/FOIL/PLUS, accesorios, ISOFRIG/EPS/PIR/ISOWALL) + matriz m lineal + recaps",
  "SKUs duplicados (GFS80, PU50MM, IF150, CAN.ISDC120); celdas #VALUE!/#REF!; filas duplicadas; typo '0mm'; IVA% inconsistente","AMARILLO/ROJO"],
 ["3","BMC crm_automatizado","1N-4kyT_uSPSVnu5tMIc6VzFIaga8FHDDEDGcclafRWg",
  "WORKBOOK CRM principal (BMC_SHEET_ID en prod). /health en vivo: tabs CRM_Operativo, Base de datos cotis de clientes, Parametros, Dashboard, Automatismos, DB, Manual, Form responses 1",
  "Drive MCP solo expuso un tab tipo payables (CONCEPTO|REFERENCIA|VENCIMIENTO|IMPORTE...); /health es autoritativo sobre tabs reales",
  "CAVEAT: Drive MCP no enumera todos los tabs de libros grandes. Nombre 'crm_automatizado' confunde. Bank account en celda REFERENCIA (sensible)","AMARILLO"],
 ["4","CALENDARIO vencimientos","1bvnbYq7MTJRpa6xEHE5m-5JcGNI9oCFke3lsJj99tdk",
  "Calendario de pagos/vencimientos de obligaciones fijas",
  "Header idéntico al tab payables de #3 (CONCEPTO|REFERENCIA|VENCIMIENTO|IMPORTE $/U$S|ESTADO|FECHA PAGO|alertas), 28 filas + SURA/DFSK",
  "CASI-DUPLICADO del tab payables de #3 (mismas filas PROSEGUR, mismo bank account). Consolidar","AMARILLO"],
 ["5","Stock E-Commerce","1egtKJAGaATLmmsJkaa2LlCv3Ah4lmNoGMNm4l0rXJQw",
  "Inventario + pricing e-commerce (debería derivar de MATRIZ #2)",
  "Tab stock principal (Shopify/Codigo/Producto/Costo/Margen/Venta/Stock) + detalle inventario + Linea Industrial",
  "SEVERO: 106 celdas #REF! (links rotos a MATRIZ) + 66 NO_HEADER + #VALUE!; columnas 'Pedido <fecha>' ad-hoc acumuladas; solapa catálogo con #2","ROJO"],
 ["6","Administrador de Cotizaciones","1Ie0KCpgWhrGaAKGAS1giLo7xpqblOUOIHEg1QbOQuu0",
  "Intake de solicitudes de cotización (con Respuesta AI + Link Presupuesto). WOLFB_ADMIN_SHEET_ID",
  "Tab quotes (Asig|Estado|Fecha|Cliente|Origen|Tel|Dir|Consulta|Respuesta AI|Link) + variante con Comentarios|Monto + bloque Borrador PDF + tab leftover Column 1/2/3",
  "Dos schemas de header para la misma lista; columna Borrador PDF duplicada; typo 'ink to folder'; tab scratch sin template","AMARILLO"],
 ["7","Pagos pendientes","1AzHhalsZKGis_oJ6J06zQeOb6uMQCsliR82VrSKUUsI",
  "Cuentas por pagar/cobrar a proveedores (ACOPIO/STOCK + DIRECTA) — RO estricto",
  "Tab ACOPIO (~35) + DIRECTA (~11) + tab export VACÍO (COTIZACION_ID...FECHA_INGRESO) + ledger histórico 2020-21 (~150) + deuda acumulada",
  "Monedas/decimales mezclados ($63955 vs USD 585,62 vs 17.281,00); ESTADO free-text; histórico mezclado con actual; tab export bridge sin poblar","AMARILLO"],
 ["8","Ventas 2026 BETA","1niLUAb9Uo1Ez4uxQTa9wxGntcUVDxXaXDL_uAQHxZXQ",
  "Workflow end-to-end lead->cotización->pedido->entrega (sucesor unificado previsto)",
  "1 tab muy ancho (38 cols) que rastrea toda la venta. Solo filas de prueba/placeholder",
  "Template no poblado (BETA); TELEFONO mismo nº en todas las filas (placeholder); muchas columnas checklist vacías","BETA"],
]

# ---------------------------------------------------------------------------
# TAB 3 — REPO ANALYSIS
# ---------------------------------------------------------------------------
repo_cols = ["Categoría","Métrica / Item","Valor","Notas / Deuda técnica"]
repo = [
 ["Escala","Archivos trackeados","2413",""],
 ["Escala","Ramas remotas","53","Brief decía 84; hubo limpieza. Solo main protegida."],
 ["Escala","PRs abiertas","45","Mayoría draft. Cluster copilot/* fix-deploy (churn de pipeline) + cursor[bot] test-coverage + dependabot."],
 ["Código","Líneas .js","92804","374 archivos"],
 ["Código",".jsx","48281","112 archivos"],
 ["Código",".mjs","15539","86 archivos"],
 ["Código","TypeScript","52 (.ts) + 376 (.tsx)","Adopción ~nula a pesar de presencia."],
 ["Docs","Markdown","136325 líneas / 1310 archivos","Docs > todo el código de la app. SPRAWL = deuda dominante."],
 ["Archivo grande","src/components/PanelinCalculadoraV3_backup.jsx","7235 líneas","Componente canónico monolítico (calculadora)."],
 ["Archivo grande","src/components/RoofPreview.jsx","4238 líneas",""],
 ["Archivo grande","server/routes/bmcDashboard.js","3603 líneas","Ruta Sheets-backed mega."],
 ["Código muerto","src/components/PanelinCalculadoraV3_legacy_inline.jsx","2354 líneas","Legacy en src/."],
 ["Código muerto","src/hooks/handoff/install-package/.../bmc-pdf-template.html.js","1247 líneas","Copia byte-idéntica del template PDF."],
 ["Deuda inline","TODO/FIXME/HACK/console.log/eslint-disable","62 total","Baja densidad (32 src + 30 server). FIXME=0."],
 ["Dockerfiles","Total","4 + .dockerignore","server/Dockerfile = CANÓNICA. 3 cloudbuild .yaml compiten -> drift."],
 ["Dockerfiles","Inconsistencia","deploy-calc-api.yml","path-filter observa Dockerfile.bmc-dashboard pero buildea server/Dockerfile."],
 ["Sub-app","shop-chat-agent","Node 18","Repo requiere Node 24 -> stack inconsistente. Dockerfile propio."],
 ["Clutter raíz","docs 2/","Duplicado verbatim de docs/","Riesgo de edición divergente."],
 ["Clutter raíz","docs.zip","133 KB","Snapshot zip commiteado."],
 ["Clutter raíz","reporte.md","739 KB","Mayor doc; reporte ad-hoc en raíz."],
 ["Clutter raíz","Obsidian /, Untitled, GEMINI.md (0b), 'El sistema de cotas...ini'","varios","Vault Obsidian + archivos huérfanos trackeados."],
 ["Sub-packages","transportista / wa / traktime / supabase","migraciones SQL","Aplicadas via npm run *:migrate."],
 ["CI","Workflows","9","ci, deploy-calc-api, deploy-vercel, smoke-prod-scheduled, knowledge-antenna(x2), drive-oauth-verify(x2), email-ingest."],
 ["Tests","Scripts en tests/","~68","Standalone node scripts (sin runner). Cubren identity, WA, agent, calc, pdf, ML, market-intel."],
 ["Org","Repos hermanos en matiasportugau-ui","24","Solo Calculadora-BMC (2026-06-04) y bmc-development-team (2026-06-03) activos. Resto stale/experimental, 0 archivados."],
]

# ---------------------------------------------------------------------------
# TAB 4 — ENDPOINTS (~226)
# ---------------------------------------------------------------------------
ep_cols = ["METHOD","PATH","Router","Auth","Dependencia de datos"]
endpoints = [
 # index.js inline
 ["GET","/capabilities","index.js","PUBLIC","compute (manifest; expone panelin_relax_dev_auth - H1)"],
 ["GET","/health","index.js","PUBLIC","Sheets probe + ML token store"],
 ["POST","/api/vitals","index.js","PUBLIC","none"],
 ["GET","/auth/ml/start","index.js","PUBLIC","ML OAuth"],
 ["GET","/auth/ml/callback","index.js","PUBLIC (state)","ML OAuth"],
 ["GET","/auth/ml/status","index.js","PUBLIC","ML token store"],
 ["GET","/ml/users/me","index.js","PUBLIC","ext: MercadoLibre"],
 ["GET","/ml/users/:id","index.js","PUBLIC","ext: MercadoLibre"],
 ["GET","/ml/listings","index.js","PUBLIC","ext: MercadoLibre"],
 ["GET","/ml/items/:id","index.js","PUBLIC","ext: MercadoLibre"],
 ["PATCH","/ml/items/:id","index.js","PUBLIC","ext: MercadoLibre (WRITE proxy - riesgo)"],
 ["POST","/ml/items/:id/description","index.js","PUBLIC","ext: MercadoLibre (WRITE proxy - riesgo)"],
 ["GET","/ml/questions","index.js","PUBLIC","ext: MercadoLibre"],
 ["GET","/ml/questions/:id","index.js","PUBLIC","ext: MercadoLibre"],
 ["POST","/ml/questions/:id/answer","index.js","PUBLIC","ext: MercadoLibre (WRITE proxy - riesgo)"],
 ["GET","/ml/orders","index.js","PUBLIC","ext: MercadoLibre"],
 ["GET","/ml/orders/:id","index.js","PUBLIC","ext: MercadoLibre"],
 ["POST","/webhooks/ml","index.js","HMAC ML","ext ML + Sheets"],
 ["GET","/webhooks/ml/events","index.js","PUBLIC","in-memory buffer (H3 - expone payloads)"],
 ["GET","/api/ml/auto-mode","index.js","PUBLIC","file flag"],
 ["POST","/api/ml/auto-mode","index.js","API_TOKEN","file flag"],
 ["GET","/webhooks/whatsapp","index.js","PUBLIC (verify)","none (Meta verify)"],
 ["POST","/webhooks/whatsapp","index.js","HMAC WA","Postgres + Sheets + agentCore"],
 ["GET","/","index.js","PUBLIC","redirect /finanzas"],
 # calc.js
 ["GET","/calc/openapi","calc.js","PUBLIC","compute"],
 ["GET","/calc/gpt-entry-point","calc.js","PUBLIC","compute"],
 ["POST","/calc/interaction-log","calc.js","PUBLIC","log file"],
 ["GET","/calc/interaction-log/list","calc.js","JWT","log files"],
 ["GET","/calc/interaction-log/file/:name","calc.js","JWT","log file"],
 ["POST","/calc/cotizar/presupuesto-libre","calc.js","PUBLIC","compute (calc)"],
 ["POST","/calc/cotizar","calc.js","PUBLIC","compute (calc)"],
 ["POST","/calc/cotizar/pdf","calc.js","JWT opt","PDF + GCS + Sheets"],
 ["GET","/calc/pdf/:id","calc.js","PUBLIC","pdfStore"],
 ["GET","/calc/cotizaciones","calc.js","JWT","quote registry"],
 ["GET","/calc/cotizaciones/:id","calc.js","JWT","quote registry"],
 ["POST","/calc/cotizaciones/:id/cancelar","calc.js","JWT","quote registry"],
 ["GET","/calc/catalogo","calc.js","PUBLIC","constants"],
 ["GET","/calc/escenarios","calc.js","PUBLIC","constants"],
 ["GET","/calc/informe","calc.js","PUBLIC","compute"],
 # legacyQuote.js
 ["GET","/ready","legacyQuote.js","PUBLIC","none"],
 ["POST","/find_products","legacyQuote.js","API_TOKEN","compute"],
 ["POST","/resolve_product","legacyQuote.js","API_TOKEN","compute"],
 ["POST","/product_price","legacyQuote.js","API_TOKEN","compute"],
 ["POST","/check_availability","legacyQuote.js","API_TOKEN","compute"],
 ["POST","/calculate_quote","legacyQuote.js","API_TOKEN","compute (calc)"],
 ["POST","/calculate_quote_v2","legacyQuote.js","API_TOKEN","compute (calc)"],
 # deepResearch.js
 ["POST","/api/research/deep","deepResearch.js","PUBLIC","ext: web search + LLM"],
 ["GET","/api/research/deep/:id","deepResearch.js","PUBLIC","job store"],
 ["POST","/api/research/deep/:id/cancel","deepResearch.js","PUBLIC","job store"],
 # agentChat.js
 ["GET","/api/agent/ai-options","agentChat.js","PUBLIC","config"],
 ["GET","/api/agent/tool-stats","agentChat.js","PUBLIC","stats"],
 ["GET","/api/agent/tools-manifest","agentChat.js","PUBLIC","compute"],
 ["POST","/api/agent/exec-tool","agentChat.js","API_TOKEN (per-tool)","loopback calc + ext"],
 ["POST","/api/agent/chat","agentChat.js","PUBLIC (rate-limited)","LLM + RAG + loopback calc"],
 # agentTraining.js (DEVMODE x27)
 ["*","/api/agent/train* (27 rutas)","agentTraining.js","DEVMODE","training KB (pgvector) + LLM"],
 # agentConversations.js (DEVMODE x6)
 ["*","/api/agent/conversations* + /stats (6 rutas)","agentConversations.js","DEVMODE","conv logs + LLM"],
 # agentVoice.js
 ["POST","/api/agent/voice/session","agentVoice.js","JWT","ext: OpenAI Realtime"],
 ["POST","/api/agent/voice/action","agentVoice.js","PUBLIC","loopback calc"],
 ["GET","/api/agent/voice/errors","agentVoice.js","JWT","in-memory"],
 ["POST","/api/agent/voice/errors/clear","agentVoice.js","JWT","in-memory"],
 ["GET","/api/agent/voice/health","agentVoice.js","JWT","config"],
 ["POST","/api/agent/transcribe","agentTranscribe.js","PUBLIC (rate-limited)","ext: OpenAI Whisper"],
 ["POST","/api/agent/feedback","agentFeedback.js","PUBLIC","store"],
 ["GET","/api/agent/feedback","agentFeedback.js","JWT","store"],
 ["GET","/api/agent/feedback/stats","agentFeedback.js","JWT","store"],
 ["GET","/api/ai-analytics/trends","aiAnalytics.js","API_TOKEN","analytics store"],
 # bmcDashboard.js (Sheets) — selección crítica
 ["GET","/api/cotizaciones","bmcDashboard.js","PUBLIC","Sheets"],
 ["POST","/api/cotizaciones","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["PATCH","/api/cotizaciones/:id","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["GET","/api/proximas-entregas","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/coordinacion-logistica","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/audit","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/pagos-pendientes","bmcDashboard.js","PUBLIC","Sheets (pagos)"],
 ["POST","/api/pagos","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["PATCH","/api/pagos/:id","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["GET","/api/metas-ventas","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/calendario-vencimientos","bmcDashboard.js","PUBLIC","Sheets (calendario)"],
 ["GET","/api/ventas","bmcDashboard.js","PUBLIC","Sheets"],
 ["POST","/api/ventas","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["GET","/api/ventas/tabs","bmcDashboard.js","PUBLIC","Sheets"],
 ["POST","/api/marcar-entregado","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["GET","/api/stock-ecommerce","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/stock-kpi","bmcDashboard.js","PUBLIC","Sheets"],
 ["PATCH","/api/stock/:codigo","bmcDashboard.js","API_TOKEN","Sheets WRITE"],
 ["GET","/api/stock/history","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/productos-maestro","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/productos-maestro/reconcile","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/productos-maestro/links","bmcDashboard.js","PUBLIC","Sheets/local"],
 ["PUT","/api/productos-maestro/links","bmcDashboard.js","PUBLIC","Sheets WRITE (sin auth - riesgo)"],
 ["POST","/api/productos-maestro/push","bmcDashboard.js","API_TOKEN","Sheets WRITE"],
 ["GET","/api/kpi-financiero","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/kpi-report","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/fiscal/bps-irae","bmcDashboard.js","PUBLIC","Sheets"],
 ["GET","/api/actualizar-precios-calculadora","bmcDashboard.js","PUBLIC","Sheets (MATRIZ CSV)"],
 ["POST","/api/matriz/push-pricing-overrides","bmcDashboard.js","API_TOKEN","Sheets WRITE"],
 ["POST","/api/crm/suggest-response","bmcDashboard.js","PUBLIC","LLM + Sheets"],
 ["POST","/api/crm/parse-email","bmcDashboard.js","PUBLIC","LLM"],
 ["POST","/api/crm/ingest-email","bmcDashboard.js","API_TOKEN","LLM + Sheets"],
 ["POST","/api/crm/parse-conversation","bmcDashboard.js","PUBLIC","LLM"],
 ["GET","/api/crm/cockpit-token","bmcDashboard.js","PUBLIC","config (DEVUELVE el token - riesgo)"],
 ["*","/api/crm/cockpit/* (13 rutas)","bmcDashboard.js","API_TOKEN","Sheets + ML/WA"],
 ["GET","/api/consultations","bmcDashboard.js","API_TOKEN","Sheets"],
 ["GET","/api/email/panelsim-summary","bmcDashboard.js","API_TOKEN","local/Sheets"],
 ["POST","/api/email/draft-outbound","bmcDashboard.js","API_TOKEN","LLM + Sheets"],
 # followups
 ["*","/api/followups* (7 rutas)","followups.js","JWT","store local"],
 # transportista
 ["GET","/api/transportista/health","transportista.js","PUBLIC","Postgres ping"],
 ["*","/api/trips* (7 rutas)","transportista.js","API_TOKEN","Postgres"],
 ["*","/api/driver/* (6 rutas)","transportista.js","driver token","Postgres + GCS"],
 # wa
 ["*","/api/wa/* config/settings/operators/rules/webhooks (12)","wa.js","WA-access","Postgres + LLM"],
 ["*","/api/wa/auth/* (magic-link, 6)","wa.js","PUBLIC (token)","Postgres + email"],
 ["GET","/api/wa/health","wa.js","PUBLIC","Postgres ping"],
 ["GET","/api/wa/config/extension","wa.js","PUBLIC (intencional)","config"],
 ["*","/api/wa/ingest|conversations|messages|suggestions|quotes|outbound|... (18)","wa.js","API_TOKEN","Postgres + WA/LLM"],
 # traktime
 ["GET","/api/traktime/health","traktime.js","PUBLIC","Postgres ping"],
 ["*","/api/traktime/* user (timer/entries/reports/clients/projects) (13)","traktime.js","JWT-user","Postgres"],
 ["*","/api/traktime/* admin (invoices/clients/projects/mirror) (18)","traktime.js","JWT-admin","Postgres + GCS + Sheets"],
 # quotes / superAgent / panelinInternal / presup / wolfboard / marketing / pdf / planInterpret
 ["GET","/api/quotes/counter","quotes.js","PUBLIC","counter atómico"],
 ["POST","/api/quotes/counter/next","quotes.js","PUBLIC","counter WRITE"],
 ["POST","/api/agent/quote-lead","superAgent.js","API_TOKEN","LLM + loopback calc"],
 ["*","/api/internal/panelin/* (4)","panelinInternal.js","API_TOKEN","dispatch tools"],
 ["POST","/api/internal/presup/run","internal/presupOrchestrator.js","PUBLIC (auth = TODO!)","orchestrator + LLM (riesgo)"],
 ["GET","/api/internal/presup/status","internal/presupOrchestrator.js","PUBLIC","in-memory"],
 ["GET","/api/internal/presup/run/example","internal/presupOrchestrator.js","PUBLIC","compute"],
 ["*","/api/wolfboard/* (7)","wolfboard.js","API_TOKEN","Sheets + loopback calc"],
 ["*","/api/marketing/* (6)","marketing.js","JWT","marketIntel ETL (Postgres)"],
 ["POST","/api/pdf/generate","pdf.js","PUBLIC","Playwright/Chromium"],
 ["GET","/api/pdf/metrics","pdf.js","PUBLIC","in-memory"],
 ["POST","/api/plan/interpret","planInterpret.js","PUBLIC (rate-limited)","LLM vision"],
 ["GET","/api/ml/search","mlSearch.js","API_TOKEN","ext MercadoLibre (cached)"],
 ["*","/api/ml/etl-run* (3)","mlEtlRun.js","API_TOKEN","marketIntel ETL (Postgres)"],
 ["*","/api/team-assist/* (3)","teamAssist.js","PUBLIC","ext OpenAI"],
 # shopify
 ["GET","/auth/shopify","shopify.js","PUBLIC","Shopify OAuth"],
 ["GET","/auth/shopify/callback","shopify.js","HMAC","Shopify OAuth"],
 ["POST","/webhooks/shopify","shopify.js","HMAC","ext Shopify"],
 ["GET","/api/shopify/products","shopify.js","API_TOKEN","ext Shopify"],
 ["GET","/api/shopify/catalog/full","shopify.js","API_TOKEN","ext Shopify"],
 ["GET","/admin/questions","shopify.js","PUBLIC","ext Shopify (H2 - sin auth)"],
 ["POST","/admin/answer","shopify.js","PUBLIC","ext Shopify WRITE (H2 - sin auth)"],
 ["GET","/admin/auto-config","shopify.js","PUBLIC","config (H2)"],
 ["POST","/admin/auto-config","shopify.js","PUBLIC","config WRITE (H2)"],
 # auth
 ["*","/api/auth/google|refresh|logout|me|me/grants (5)","authGoogle.js","PUBLIC/JWT","Google OAuth + Postgres"],
 ["*","/api/auth/mfa/* (4)","authMfa.js","JWT/PUBLIC(challenge)","Postgres TOTP"],
 # identity
 ["*","/api/me/* + access-requests + threads (25)","identityMe.js","JWT-user (admin subset)","Postgres (+Sheets admin)"],
 ["*","/api/admin/users* role-grants/suspend (8)","identityAdmin.js","JWT-admin","Postgres identity"],
 ["*","/api/admin/analytics/* (5)","identityAnalytics.js","JWT-admin","Postgres activity"],
 ["*","/api/clientes/customers|followups (4)","clientes/*.js","JWT + grant","Sheets clientes"],
 ["*","/api/admin/export + /api/me/quotes/:id/export.* (5)","quoteExport.js","JWT","Postgres + PDF"],
 # tasks
 ["*","/api/tasks/* lists+tasks (9)","tasks.js","JWT-user","ext Google Tasks"],
 ["*","/auth/tasks/* (4)","tasksOAuth.js","JWT/PUBLIC(callback)","Google OAuth + Postgres"],
 ["POST","/sync/google-tasks/pull","tasksSync.js","HMAC (x-sync-signature)","ext Google Tasks + Postgres"],
]

# ---------------------------------------------------------------------------
# TAB 5 — SECURITY FINDINGS
# ---------------------------------------------------------------------------
sec_cols = ["ID","Severidad","Categoría","Ubicación (file:line)","Descripción (redactada)","Recomendación"]
security = [
 ["C1","CRITICAL","PII en archivos trackeados",".accessible-base/crm_operativo.json",
  "Snapshot CRM commiteado: 297 filas de leads (telefonos reales) + spreadsheet_id en _meta. NO está en .gitignore. Confirmado: tracked, 151KB, rows=297.",
  "Quitar de git history (git filter-repo), agregar /.accessible-base/ a .gitignore. Evaluar reporte bajo Ley 18.331 (datos personales UY)."],
 ["C2","CRITICAL","Secreto default hardcodeado",".env.example:312",
  "WA_JWT_SECRET = literal 'change-me-...' commiteado. Quien copie .env.example sin cambiarlo firma JWT con secreto público (tokens forjables).",
  "Placeholder vacío + hard-fail en waOperatorAuth.getJwtSecret() si detecta 'change-me'."],
 ["C3","CRITICAL","API token en bundle Vite","src/utils/teamAssistApi.js:5; AgentAdminModule:28; PricingEditor:57; BmcLogisticaApp:1468",
  "VITE_API_AUTH_TOKEN / VITE_BMC_API_AUTH_TOKEN leídos de import.meta.env -> si están seteados en el build, el bearer del server viaja en el JS público.",
  "Usar cookie de sesión / credencial frontend corta. CI check que falle si esas vars están en el build Vercel."],
 ["H1","HIGH","Flag de bypass expuesto","server/agentCapabilitiesManifest.js:118 + index.js:179",
  "GET /capabilities (público) devuelve panelin_relax_dev_auth. Si true en prod, revela que devMode saltea auth.",
  "Quitar el flag de la respuesta pública."],
 ["H2","HIGH","Rutas Shopify admin sin auth","server/routes/shopify.js:511,534,566,581",
  "/admin/questions, /admin/answer, /admin/auto-config (GET+POST) sin middleware de auth. Cualquiera lee/escribe.",
  "Aplicar requireApiAuth a las 4 rutas /admin/*."],
 ["H3","HIGH","Log de webhook ML sin auth","server/index.js:559-562",
  "GET /webhooks/ml/events devuelve el ring buffer de 250 payloads ML (incl x-signature, seller data) sin auth.",
  "Bearer apiAuthToken antes de devolver eventos."],
 ["H4","HIGH","WA JWT secret cae a API_AUTH_TOKEN","server/lib/waOperatorAuth.js:42",
  "getJwtSecret() usa WA_JWT_SECRET || API_AUTH_TOKEN. Si WA_JWT_SECRET falta, firma con el mismo secreto de writes CRM.",
  "Quitar fallback a API_AUTH_TOKEN; exigir WA_JWT_SECRET independiente."],
 ["H5","HIGH","Google OAuth Client ID commiteado",".env.example:364",
  "GOOGLE_OAUTH_CLIENT_ID real (...apps.googleusercontent.com) en el example. Facilita phishing + revela nº proyecto.",
  "Reemplazar por placeholder."],
 ["W1","MEDIUM","Sheet IDs hardcodeados como fallback","server/config.js:68,84",
  "wolfbAdminSheetId (1Ie0KCpg...) y bmcMatrizSheetId (1oDMkBgWx...) hardcodeados. Deploy mal configurado escribe en sheets prod en silencio.",
  "Fallback '' y fallar 503 si falta env var."],
 ["W2","MEDIUM","ML Client ID hardcodeado","server/config.js:24",
  "mlClientId fallback '742811153438318'. ID público pero mismo mal patrón.","Fallback ''."],
 ["W3","MEDIUM","Wildcard chrome-extension en CORS","server/index.js:99-100",
  "Cualquier chrome-extension:// permitido. Extensión maliciosa local podría hacer requests credenciadas.",
  "Allowlist de IDs de extensión específicos."],
 ["W4","MEDIUM","Shopify webhook HMAC sin warning si falta secret","server/routes/shopify.js:54-56",
  "Bloquea correctamente si secret vacío, pero sin log (a diferencia de WhatsApp). Difícil detectar misconfig.",
  "Agregar warning log cuando shopifyWebhookSecret vacío."],
 ["W5","MEDIUM","GOOGLE_APPLICATION_CREDENTIALS apunta dentro del repo",".env.example:52",
  "Default = docs/.../service-account.json (dentro del árbol). Frágil: riesgo de commit accidental (gitignore lo cubre).",
  "Default vacío + comentario apuntando a path absoluto fuera del repo."],
 ["W6","MEDIUM","Más archivos de datos trackeados",".accessible-base/admin_cotizaciones.json, matriz_precios.json, manifest.json",
  "Enumeran spreadsheet_id en vivo y estructura de negocio.","Excluir todo .accessible-base/ de git."],
 ["W7","MEDIUM","Service account email hardcodeado","scripts/ensure-panelsim-sheets-env.sh:9",
  "bmc-dashboard-sheets@chatbot-bmc-live.iam... como default. Enumera proyecto + SA.","Mover a env var sin fallback."],
 ["W8","LOW","URL Cloud Run prod en comentarios",".env.example:81,108,251",
  "URL panelin-calc-642127786762... x3 en comentarios. Pública pero + nº proyecto facilita recon.",
  "Reemplazar por $PUBLIC_BASE_URL en el example."],
]
sec_passed = [
 "Sin secretos reales en código (no sk-/AIza/AKIA/PEM fuera de placeholders).",
 ".env NO trackeado; nunca commiteado (git log -- *.env vacío).",
 "service-account.json no trackeado; cubierto por .gitignore.",
 "Webhooks WhatsApp/ML/Shopify/Tasks con HMAC + timingSafeEqual; raw body preservado.",
 "IDENTITY_JWT_SECRET con hard-fail (>=32 chars, != WA_JWT_SECRET).",
 "tokenStore AES-256-GCM con IV random + auth tag; key validada.",
 "CORS restringido a allowlist en prod (sin '*'); credentials con origin allowlist.",
 "Rate limiting en chat/exec-tool/transcribe; chat no autenticado no llama tools auth-gated.",
 "TOTP KEK (MFA_KEK_HEX) separado del JWT secret.",
 "Headers de seguridad (X-Frame-Options, nosniff, HSTS, CSP) en vercel.json + server.",
]

# ---------------------------------------------------------------------------
# TAB 6 — CRITICAL QUESTIONS (P1-P5)
# ---------------------------------------------------------------------------
q_cols = ["ID","Pregunta de decisión","Opciones","Estado prior (2026-04-23)","Evidencia NUEVA en vivo (2026-06-07)","Respuesta recomendada","Bloquea"]
questions = [
 ["P1","Modelo de deploy canónico de API + frontend",
  "(a) 2 servicios GH Actions; (b) full-stack Cloud Build GCR legacy; (c) Vercel frontend + Cloud Run panelin-calc API; (d) otro",
  "ABIERTA — 3 modelos coexisten",
  "vercel.json -> Cloud Run q74 API; /capabilities public_base_url = q74; CI activo = deploy-calc-api.yml con server/Dockerfile. La realidad operativa = (c).",
  "(c). Borrar/archivar cloudbuild.yaml, cloudbuild-frontend.yaml, Dockerfile.bmc-dashboard, Dockerfile(root) o documentarlos como NO-CI. Arreglar path-filter de deploy-calc-api.yml.",
  "Roadmap deploy/rollback"],
 ["P2","Master de precios activo",
  "(a) Sheet MATRIZ via /api/actualizar-precios; (b) constants.js hardcoded; (c) ambos con drift",
  "AMARILLA — 2 caminos paralelos",
  "PROJECT-STATE 2026-06-02/04: MATRIZ mapping desplegado a prod, /api/actualizar-precios 200; constants.js parcialmente bakeado. MATRIZ tiene SKUs dup + #REF!.",
  "(a) como fuente, constants.js solo fallback. PERO primero limpiar MATRIZ (dedupe SKUs, arreglar #REF!) antes de tratarla como single source of truth.",
  "Correctitud de cotizaciones"],
 ["P3","Schema CRM activo en prod (BMC_SHEET_SCHEMA)",
  "(a) CRM_Operativo; (b) Master_Cotizaciones; (c) ambos; (d) verificar",
  "ABIERTA — requería gcloud",
  "RESUELTA: /health en vivo lista tab CRM_Operativo (missing=[]). Schema activo = CRM_Operativo.",
  "(a) CRM_Operativo (confirmado en vivo). Retirar/archivar el path Master_Cotizaciones si no se usa.",
  "Integración CRM"],
 ["P4","Estado de repos hermanos",
  "(a) solo 3; (b) todos vivos; (c) mayoría experimentales, operación solo en calculadora-bmc; (d) listar",
  "ABIERTA — requería gh",
  "RESUELTA: 24 repos en matiasportugau-ui. Solo Calculadora-BMC (2026-06-04) y bmc-development-team (2026-06-03) activos. omnicrm-sync existe. Resto stale (Jan-Mar) / 0 archivados.",
  "(c). Archivar formalmente los repos GPT/chatbot experimentales para reducir superficie (forks viejos = riesgo de credenciales/CI).",
  "Higiene / seguridad org"],
 ["P5","Master Apps Script (.gs): repo == lo que corre?",
  "(a) sí, sync manual clasp; (b) no, verdad en Apps Script bound; (c) divergencia conocida; (d) otro",
  "ABIERTA — sin clasp/Apps Script access",
  "SIGUE ABIERTA. .gs usan getActiveSpreadsheet() (container-bound). Sin pipeline. CalendarioRecordatorio usa WORKBOOK1_ID de ScriptProperties (no en repo).",
  "(c) probable. Adoptar clasp + repo como verdad, o documentar explícitamente que el bound script es el master y los .gs son snapshots.",
  "Confiabilidad dashboards/automatizaciones"],
]

# ---------------------------------------------------------------------------
# TAB 7 — RECOMMENDATIONS
# ---------------------------------------------------------------------------
rec_cols = ["Prioridad","Área","Acción","Razón","Esfuerzo"]
recs = [
 ["P0","Seguridad/PII","Purgar .accessible-base/ del repo + git history; .gitignore","C1: 297 leads + telefonos commiteados. Posible obligación Ley 18.331.","M"],
 ["P0","Seguridad","Quitar default WA_JWT_SECRET + hard-fail (C2); sacar API token del bundle Vite (C3)","Tokens forjables / leak de bearer de server.","S-M"],
 ["P0","Seguridad","Auth en writes Sheets públicos (POST/PATCH /api/cotizaciones,/pagos,/ventas,/marcar-entregado, PUT productos-maestro/links) + /api/crm/cockpit-token + /admin/* (H2) + /webhooks/ml/events (H3) + /api/internal/presup/*","Mutación de datos de producción sin auth.","M"],
 ["P1","Infra","Decidir P1 y borrar los 2 modelos de deploy muertos + arreglar path-filter deploy-calc-api.yml","Drift de deploy: imagen activa depende de quién desplegó último.","S"],
 ["P1","Datos/Precios","Limpiar MATRIZ #2: dedupe SKUs (GFS80,PU50MM,IF150,CAN.ISDC120), arreglar #REF!/#VALUE!, normalizar IVA%","Es la fuente de precios de la calc (P2). Datos sucios -> cotizaciones erróneas.","M"],
 ["P1","Datos","Arreglar 106 #REF! en Stock E-Commerce #5 (links rotos a MATRIZ)","Inventario/pricing e-commerce roto.","M"],
 ["P1","Config","Setear ML_CLIENT_SECRET en prod (/health missingConfig) o documentar que ML OAuth está intencionalmente off","ML OAuth no funcional en prod.","S"],
 ["P2","Datos","Consolidar #3 crm_automatizado ~ #4 CALENDARIO (payables duplicados) en un libro; renombrar #3 (no es CRM)","Dos copias del mismo dataset de vencimientos.","S"],
 ["P2","Datos","Unificar lifecycle lead->venta: decidir si #8 'Ventas 2026 BETA' reemplaza a #1+#6; poblar tab export vacío de #7","3 sheets con columnas solapadas para el mismo flujo.","M-L"],
 ["P2","Repo","Limpieza de clutter: borrar docs 2/, docs.zip, reporte.md, Obsidian/, Untitled, GEMINI.md; mover reportes raíz a docs/","136K líneas md; duplicados divergentes.","S"],
 ["P2","Repo","Consolidar 45 PRs abiertas; cerrar ramas stale; resolver código muerto (*_legacy_inline.jsx, template PDF duplicado)","Deuda de integración alta.","M"],
 ["P3","Repo","Alinear shop-chat-agent a Node 24 o aislarlo como repo propio","Stack inconsistente (Node 18 vs 24).","S"],
 ["P3","Org","Archivar repos hermanos experimentales (GPT-*, chatbot-*, aistudioPanelin, etc.)","Reduce superficie de credenciales/CI muertos (P4).","S"],
 ["P3","Apps Script","Adoptar clasp (repo = verdad) o documentar bound script como master (P5)","Sin forma de saber qué .gs corre.","M"],
 ["P3","Seguridad","Aplicar W1-W8 (fallbacks vacíos para sheet/ML IDs, allowlist extension CORS, warnings HMAC)","Endurecimiento defensivo.","S"],
]

# ---------------------------------------------------------------------------
# WRITE CSVs
# ---------------------------------------------------------------------------
def write_csv(name, cols, rows):
    p = os.path.join(DATA, name)
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(cols); w.writerows(rows)
    return p

tabs = [
    ("01_Infraestructura", infra_cols, infra),
    ("02_Sheets_Analysis", sheets_cols, sheets),
    ("03_Repo_Analysis", repo_cols, repo),
    ("04_Endpoints", ep_cols, endpoints),
    ("05_Security_Findings", sec_cols, security),
    ("06_Critical_Questions", q_cols, questions),
    ("07_Recommendations", rec_cols, recs),
]
for name, cols, rows in tabs:
    write_csv(name + ".csv", cols, rows)

# ---------------------------------------------------------------------------
# WRITE XLSX
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="E5E7EB")]*4)
SEV = {
    "CRITICAL": "FCA5A5", "HIGH": "FDBA74", "MEDIUM": "FDE68A", "LOW": "BBF7D0",
    "P0": "FCA5A5", "P1": "FDBA74", "P2": "FDE68A", "P3": "BBF7D0",
}
WIDTHS = {
    "01_Infraestructura":[14,30,42,26,16,60],
    "02_Sheets_Analysis":[4,26,46,40,46,52,12],
    "03_Repo_Analysis":[16,40,16,60],
    "04_Endpoints":[8,52,22,20,46],
    "05_Security_Findings":[6,12,26,40,60,52],
    "06_Critical_Questions":[5,34,46,28,52,52,22],
    "07_Recommendations":[10,16,60,52,8],
}

# Overview sheet first
ws = wb.active; ws.title = "00_Overview"
ws.sheet_view.showGridLines = False
ov_rows = [
    ["BMC URUGUAY — AUDITORÍA TÉCNICA MAESTRO", ""],
    ["", ""],
    ["Organización", overview["org"]],
    ["Sistema", overview["system"]],
    ["Repo", overview["repo"]],
    ["Fecha auditoría", overview["audit_date"]],
    ["Construye sobre", overview["builds_on"]],
    ["", ""],
    ["PRODUCCIÓN EN VIVO", ""],
    ["Cloud Run", overview["live_prod"]["cloud_run_url"]],
    ["/health", overview["live_prod"]["health"]],
    ["gitSha desplegado", LIVE_SHA + "  (== main local)"],
    ["missingConfig", ", ".join(overview["live_prod"]["missing_config"])],
    ["", ""],
    ["NÚMEROS CLAVE", ""],
]
for k, v in overview["headline_numbers"].items():
    ov_rows.append([k.replace("_"," "), v])
ov_rows.append(["", ""])
ov_rows.append(["DELTAS vs AUDITORÍA PREVIA", ""])
for d in overview["deltas_vs_prior"]:
    ov_rows.append(["", d])
for r in ov_rows:
    ws.append(r)
# Style: title row + any section-header row (matched by label, not by fixed coords).
SECTION_HEADERS = {"PRODUCCIÓN EN VIVO", "NÚMEROS CLAVE", "DELTAS vs AUDITORÍA PREVIA"}
ws["A1"].font = Font(bold=True, size=16, color="1F2937")
for ridx in range(2, ws.max_row + 1):
    label = str(ws.cell(row=ridx, column=1).value or "").strip()
    if label in SECTION_HEADERS:
        ws.cell(row=ridx, column=1).font = Font(bold=True, size=12, color="B45309")
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 95
for row in ws.iter_rows():
    for c in row:
        c.alignment = WRAP

def add_tab(title, cols, rows):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws.append(cols)
    for c in ws[1]:
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = WRAP; c.border = THIN
    for r in rows:
        ws.append(r)
    # widths
    w = WIDTHS.get(title)
    for i, col in enumerate(cols):
        letter = get_column_letter(i+1)
        ws.column_dimensions[letter].width = (w[i] if w and i < len(w) else 24)
    # severity coloring (col B for security/recs)
    sev_idx = None
    if title == "05_Security_Findings": sev_idx = 2
    if title == "07_Recommendations": sev_idx = 1
    for ridx in range(2, ws.max_row+1):
        for c in ws[ridx]:
            c.alignment = WRAP; c.border = THIN
        if sev_idx:
            key = str(ws.cell(row=ridx, column=sev_idx).value or "").strip().upper()
            fill = SEV.get(key)
            if fill:
                ws.cell(row=ridx, column=sev_idx).fill = PatternFill("solid", fgColor=fill)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

for name, cols, rows in tabs:
    add_tab(name, cols, rows)

xlsx_path = os.path.join(HERE, "BMC-MAESTRO-AUDIT.xlsx")
wb.save(xlsx_path)

# ---------------------------------------------------------------------------
# WRITE JSON metadata
# ---------------------------------------------------------------------------
def to_objs(cols, rows):
    return [dict(zip(cols, r)) for r in rows]

meta = {
    "generated_at": AUDIT_DATE,  # deterministic: audit date, not wall-clock (avoids diff churn)
    "overview": overview,
    "infrastructure": to_objs(infra_cols, infra),
    "sheets_analysis": to_objs(sheets_cols, sheets),
    "repo_analysis": to_objs(repo_cols, repo),
    "endpoints": to_objs(ep_cols, endpoints),
    "security_findings": to_objs(sec_cols, security),
    "security_passed_checks": sec_passed,
    "critical_questions": to_objs(q_cols, questions),
    "recommendations": to_objs(rec_cols, recs),
}
json_path = os.path.join(HERE, "audit-metadata.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

# ---------------------------------------------------------------------------
# WRITE HTML dashboard
# ---------------------------------------------------------------------------
def esc(s):
    return (str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"))

def html_table(cols, rows, sev_idx=None):
    h = "<table><thead><tr>" + "".join(f"<th>{esc(c)}</th>" for c in cols) + "</tr></thead><tbody>"
    for r in rows:
        h += "<tr>"
        for i, c in enumerate(r):
            cls = ""
            if sev_idx is not None and i == sev_idx:
                key = str(c).strip().upper()
                cls = f' class="sev sev-{key}"'
            h += f"<td{cls}>{esc(c)}</td>"
        h += "</tr>"
    return h + "</tbody></table>"

hn = overview["headline_numbers"]
cards = [
    ("Endpoints API", hn["api_endpoints"], f"{hn['public_endpoints']} públicos"),
    ("Ramas / PRs", f"{hn['remote_branches']} / {hn['open_prs']}", "deuda integración"),
    ("Sheets auditadas", hn["google_sheets_audited"], "8/8 en vivo"),
    ("Repos en org", hn["sibling_repos_in_org"], "2 activos"),
    ("Hallazgos seg.", hn["security_findings"], f"{hn['critical_findings']} críticos"),
    ("Archivos repo", hn["tracked_files"], "v3.1.5"),
]
card_html = "".join(
    f'<div class="card"><div class="k">{esc(t)}</div><div class="v">{esc(v)}</div><div class="s">{esc(s)}</div></div>'
    for t,v,s in cards)

html = f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>BMC Uruguay — Audit MAESTRO {AUDIT_DATE}</title>
<style>
:root{{--bg:#faf9f7;--ink:#1f2937;--mut:#6b7280;--line:#e5e7eb;--accent:#b45309;}}
*{{box-sizing:border-box}} body{{margin:0;background:var(--bg);color:var(--ink);
font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;line-height:1.5}}
.wrap{{max-width:1180px;margin:0 auto;padding:32px 20px 80px}}
h1{{font-family:Georgia,'Times New Roman',serif;font-weight:600;font-size:30px;margin:0 0 4px}}
.sub{{color:var(--mut);margin:0 0 24px}}
h2{{font-family:Georgia,serif;font-size:21px;margin:38px 0 12px;border-bottom:2px solid var(--line);padding-bottom:6px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin:18px 0}}
.card{{background:#fff;border:1px solid var(--line);border-radius:12px;padding:14px 16px}}
.card .k{{font-size:12px;color:var(--mut);text-transform:uppercase;letter-spacing:.04em}}
.card .v{{font-size:26px;font-weight:700;margin:4px 0}}
.card .s{{font-size:12px;color:var(--accent)}}
.banner{{background:#fff7ed;border:1px solid #fed7aa;border-left:4px solid var(--accent);
border-radius:10px;padding:14px 16px;margin:14px 0}}
.banner b{{color:#9a3412}}
table{{width:100%;border-collapse:collapse;background:#fff;border:1px solid var(--line);
border-radius:10px;overflow:hidden;font-size:13px;margin:8px 0}}
th{{background:#1f2937;color:#fff;text-align:left;padding:9px 10px;font-weight:600}}
td{{padding:8px 10px;border-top:1px solid var(--line);vertical-align:top}}
tr:nth-child(even) td{{background:#fcfcfb}}
.sev{{font-weight:700;text-align:center}}
.sev-CRITICAL,.sev-P0{{background:#fca5a5}} .sev-HIGH,.sev-P1{{background:#fdba74}}
.sev-MEDIUM,.sev-P2{{background:#fde68a}} .sev-LOW,.sev-P3{{background:#bbf7d0}}
ul{{margin:6px 0 6px 18px}} li{{margin:3px 0}}
.foot{{color:var(--mut);font-size:12px;margin-top:40px;border-top:1px solid var(--line);padding-top:14px}}
.pill{{display:inline-block;background:#ecfdf5;border:1px solid #a7f3d0;color:#065f46;
border-radius:999px;padding:2px 10px;font-size:12px;margin-right:6px}}
</style></head><body><div class="wrap">
<h1>BMC Uruguay — Auditoría Técnica MAESTRO</h1>
<p class="sub">{esc(overview['system'])} · repo <code>{esc(overview['repo'])}</code> · {AUDIT_DATE}</p>
<div class="banner"><b>Producción en vivo:</b> Cloud Run <code>panelin-calc</code> HTTP 200 ·
gitSha desplegado <code>3153b11</code> = <code>main</code> local · hasSheets=true, hasTokens=true ·
<b>missingConfig: ML_CLIENT_SECRET</b>.</div>
<div class="cards">{card_html}</div>

<h2>Deltas vs auditoría previa ({PRIOR_DATE})</h2>
<ul>{''.join(f'<li>{esc(d)}</li>' for d in overview['deltas_vs_prior'])}</ul>

<h2>Preguntas críticas de decisión (P1–P5)</h2>
{html_table(q_cols, questions)}
<p><span class="pill">P3 resuelta en vivo</span><span class="pill">P4 resuelta en vivo</span>
P1 y P2 con respuesta recomendada; P5 sigue abierta.</p>

<h2>Hallazgos de seguridad ({hn['critical_findings']} críticos)</h2>
{html_table(sec_cols, security, sev_idx=1)}
<details><summary>Checks aprobados ({len(sec_passed)})</summary><ul>{''.join(f'<li>{esc(x)}</li>' for x in sec_passed)}</ul></details>

<h2>Análisis de Sheets (8 libros)</h2>
{html_table(sheets_cols, sheets)}

<h2>Infraestructura</h2>
{html_table(infra_cols, infra)}

<h2>Análisis del repositorio</h2>
{html_table(repo_cols, repo)}

<h2>Recomendaciones priorizadas</h2>
{html_table(rec_cols, recs, sev_idx=0)}

<h2>Endpoints API (resumen, ~226 totales)</h2>
{html_table(ep_cols, endpoints)}

<div class="foot">Generado por audit/build_maestro.py · auditoría read-only · sin secretos ni PII de clientes.
Acompaña: BMC-MAESTRO-AUDIT.xlsx (subir a Drive → Google Sheet), audit-metadata.json, MAESTRO_data/*.csv.</div>
</div></body></html>"""

html_path = os.path.join(HERE, "dashboard.html")
with open(html_path, "w", encoding="utf-8") as f:
    f.write(html)

print("OK — generated:")
for p in [xlsx_path, json_path, html_path]:
    print("  ", os.path.relpath(p, HERE), f"({os.path.getsize(p)} bytes)")
print("   CSVs:", len(tabs), "in MAESTRO_data/")
print("Endpoints rows:", len(endpoints), "| Security:", len(security), "| Sheets:", len(sheets), "| Recs:", len(recs))
